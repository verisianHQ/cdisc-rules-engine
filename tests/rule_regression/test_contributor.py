import json
import openpyxl as op
import os
import pytest
import yaml
from pathlib import Path
from typing import Dict, List

from cdisc_rules_engine.data_service.postgresql_data_service import PostgresQLDataService
from rule_regression.regression import (
    TEST_CACHE_PATH,
    sharepoint_xlsx_to_test_datasets,
    process_test_case_dataset,
)

SCRIPT_DIR = Path(os.path.dirname(__file__))
CONTRIBUTOR_RULES_PATH = Path(os.getenv("CONTRIBUTOR_REPO_PATH")).resolve() / "rules"


def get_contributor_test_cases() -> List[Dict]:
    cases = []
    if not CONTRIBUTOR_RULES_PATH.exists() or not CONTRIBUTOR_RULES_PATH.is_dir():
        print(f"Warning: Contributor rules path not found at {CONTRIBUTOR_RULES_PATH}")
        return cases

    for rule_dir in sorted(CONTRIBUTOR_RULES_PATH.iterdir()):
        if not rule_dir.is_dir() or not (rule_dir.name.startswith("CORE-")):
            continue

        rule_id = rule_dir.name
        rule_ymls = list(rule_dir.glob("[!~]*.yml"))

        if not rule_ymls:
            continue
        rule_yaml = rule_ymls[0]

        for test_type in ["positive", "negative"]:
            type_dir = rule_dir / test_type
            if not type_dir.exists():
                continue

            for case_dir in sorted(type_dir.iterdir()):
                if case_dir.is_dir():
                    data_dir = case_dir / "data"
                    excels = list(data_dir.glob("[!~]*.xls*"))
                    if excels:
                        cases.append(
                            {
                                "rule_id": rule_id,
                                "rule_yaml": str(rule_yaml),
                                "test_type": test_type,
                                "case_id": case_dir.name,
                                "data_path": str(data_dir),
                                "excel_file": str(excels[0]),
                                "rule_dir": str(rule_dir),
                            }
                        )
    return cases


def _read_library_specs(excel_path: str) -> dict:
    wb = op.load_workbook(excel_path, data_only=True, read_only=True)
    if "Library" not in wb.sheetnames:
        raise ValueError(f"Library sheet missing in {excel_path}")

    ws = wb["Library"]
    rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    wb.close()

    if not rows or rows[0][0] is None:
        raise ValueError(f"Library sheet in {excel_path} is missing standard/version data")

    standard = str(rows[0][0]).strip()
    version = str(rows[0][1]).strip().replace("-", ".")

    return {
        "standard": standard,
        "standard_version": version,
        "standard_substandard": None,
        "define_xml_version": None,
    }


@pytest.fixture(scope="session")
def contributor_results(pytestconfig):
    results = {}
    yield results
    output_path = Path(pytestconfig.rootpath) / "tests" / "resources" / "rules" / "contributor_results.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=4, ensure_ascii=False)


pytest_cases = get_contributor_test_cases()


@pytest.mark.parametrize(
    "case", pytest_cases, ids=[f"{c['rule_id']}-{c['test_type']}-{c['case_id']}" for c in pytest_cases]
)
def test_contributor_rule_case(case, contributor_results):
    """
    Executes a single test case from the contributor repo.
    """
    rule_id = case["rule_id"]
    test_type = case["test_type"]
    case_id = case["case_id"]

    if rule_id not in contributor_results:
        contributor_results[rule_id] = {"positive": {}, "negative": {}}

    with open(case["rule_yaml"], "r", encoding="utf-8") as f:
        rule = yaml.safe_load(f)

    ig_specs = _read_library_specs(case["excel_file"])

    data_path_obj = Path(case["data_path"])
    case_define_path = data_path_obj / "define.xml"
    rule_define_path = Path(case["rule_dir"]) / "define.xml"

    define_xml_path = None
    if case_define_path.exists():
        define_xml_path = str(case_define_path)
    elif rule_define_path.exists():
        define_xml_path = str(rule_define_path)

    data_service = PostgresQLDataService.instance(use_pgserver=True, cache_path=TEST_CACHE_PATH)

    if define_xml_path:
        data_service._update_define_xml_path(define_xml_path)

    test_datasets = sharepoint_xlsx_to_test_datasets(case["excel_file"])

    regression_errors = {}

    contributor_results[rule_id][test_type][case_id] = regression_errors

    process_test_case_dataset(
        regression_errors=regression_errors,
        define_xml_file_path=define_xml_path,
        data_test_datasets=test_datasets,
        ig_specs=ig_specs,
        rule=rule,
        test_case_folder_path=case["data_path"],
        cur_core_id=rule_id,
        use_pgserver=True,
        data_service=data_service,
    )

    assert (
        regression_errors.get("datasets_import_sql") == "SUCCESS"
    ), f"Dataset import failed: {
        regression_errors.get('datasets_conversion') or regression_errors.get('datasets_import_sql')
    }"

    assert (
        regression_errors.get("sql_overall_result") != "execution_error"
    ), f"Engine crashed during SQL execution: {json.dumps(regression_errors.get('results_sql', []), indent=2)}"

    results_sql = regression_errors.get("results_sql", [])
    total_errors = sum(len(ds.get("errors", [])) for ds in results_sql)

    if test_type == "positive":
        assert total_errors == 0, f"Expected 0 errors for a positive test, but got {total_errors}. "
        f"Details:\n{json.dumps(results_sql, indent=2)}"
    else:
        assert (
            total_errors > 0
        ), "Expected > 0 errors for a negative test, but got 0. Rule may not have triggered properly."
