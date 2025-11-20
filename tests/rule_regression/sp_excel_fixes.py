import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import os

import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


XLSX_FILE_EXTENSIONS = [".xlsx"]


def add_libraries_sheet(xlsx_path: str, sheet_name: str = "Library"):
    df = pd.DataFrame({"Product": ["sdtmig"], "Version": ["3-3"]})
    xlsx_path = Path(xlsx_path)

    if xlsx_path.exists():
        # Check for existing sheet without locking the file for writing
        wb = load_workbook(xlsx_path, read_only=True)
        try:
            if sheet_name in wb.sheetnames:
                raise FileExistsError(f"Sheet '{sheet_name}' already exists in '{xlsx_path}'.")
        finally:
            wb.close()
        mode = "a"

        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode=mode) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        raise FileNotFoundError(f"File '{xlsx_path}' not found.")


def fix_missing_libsheet(rtype: str, in_path: str):
    file_path = f"{in_path}/___{rtype}_missing_lib.txt"
    path = Path(file_path)
    if not path.is_absolute():
        raise ValueError(f"Expected absolute path, got: {file_path}")
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    excels_to_fix = []
    print(path)
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            excels_to_fix.append(f"{in_path}/{rtype}/{line.rstrip()}")

    for ex_path in excels_to_fix:
        add_libraries_sheet(ex_path)


sp_path = os.path.expanduser("~") + "/data/CORE/CDISC_Sharepoint_dump_20250916"
fix_missing_libsheet("script_testing", sp_path)
# fix_missing_libsheet("SDTMIG", sp_path)
# fix_missing_libsheet("ADAMIG", sp_path)
# fix_missing_libsheet("FDA Business Rules", sp_path)
# fix_missing_libsheet("FDA Validator Rules", sp_path)
